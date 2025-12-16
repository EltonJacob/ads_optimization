"""File upload validation and utilities."""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import UploadFile

logger = logging.getLogger(__name__)

# Supported file types
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

# Required columns for spreadsheet import
# These are the minimum columns we need to import keyword performance data
REQUIRED_COLUMNS = {
    "keyword",
    "impressions",
    "clicks",
    "spend",
    "sales",
    "orders",
}

# Optional but recommended columns
OPTIONAL_COLUMNS = {
    "date",
    "keyword_id",
    "keywordid",
    "match type",
    "match_type",
    "campaign",
    "campaign name",
    "ad group",
    "ad group name",
    "state",
    "bid",
    "cost",
    "conversions",
}


def validate_file_type(filename: str) -> tuple[bool, str]:
    """Validate file type is supported.

    Args:
        filename: Name of the uploaded file

    Returns:
        Tuple of (is_valid, error_message)
    """
    file_ext = Path(filename).suffix.lower()

    if file_ext not in ALLOWED_EXTENSIONS:
        return False, f"Unsupported file type: {file_ext}. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"

    return True, ""


def validate_file_size(size_bytes: int) -> tuple[bool, str]:
    """Validate file size is within limits.

    Args:
        size_bytes: File size in bytes

    Returns:
        Tuple of (is_valid, error_message)
    """
    if size_bytes > MAX_FILE_SIZE:
        max_mb = MAX_FILE_SIZE / (1024 * 1024)
        actual_mb = size_bytes / (1024 * 1024)
        return False, f"File too large: {actual_mb:.1f}MB. Maximum allowed: {max_mb:.0f}MB"

    if size_bytes == 0:
        return False, "File is empty"

    return True, ""


def generate_upload_id() -> str:
    """Generate a unique upload ID.

    Returns:
        Unique upload ID in format: upload_YYYYMMDD_HHMMSS_microseconds
    """
    now = datetime.now()
    return f"upload_{now.strftime('%Y%m%d_%H%M%S')}_{now.microsecond}"


def get_upload_path(upload_id: str, filename: str, profile_id: str) -> Path:
    """Get the storage path for an uploaded file.

    Args:
        upload_id: Unique upload identifier
        filename: Original filename
        profile_id: Amazon Ads profile ID

    Returns:
        Path object for the upload location
    """
    # Create directory structure: data/uploads/{profile_id}/{upload_id}_{filename}
    base_dir = Path("data/uploads") / profile_id
    base_dir.mkdir(parents=True, exist_ok=True)

    # Preserve file extension
    file_ext = Path(filename).suffix
    safe_filename = f"{upload_id}{file_ext}"

    return base_dir / safe_filename


async def save_upload_file(upload_file: UploadFile, destination: Path) -> int:
    """Save uploaded file to destination.

    Args:
        upload_file: FastAPI UploadFile object
        destination: Destination path

    Returns:
        Number of bytes written
    """
    bytes_written = 0

    # Ensure parent directory exists
    destination.parent.mkdir(parents=True, exist_ok=True)

    # Write file in chunks to handle large files
    with open(destination, "wb") as f:
        while chunk := await upload_file.read(8192):  # 8KB chunks
            bytes_written += len(chunk)
            f.write(chunk)

    return bytes_written


def validate_csv_columns(file_path: Path) -> tuple[list[str], list[str]]:
    """Validate CSV file has required columns.

    Args:
        file_path: Path to the CSV file

    Returns:
        Tuple of (detected_columns, missing_columns)
    """
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                return [], list(REQUIRED_COLUMNS)

            # Normalize column names to lowercase for comparison
            detected_columns = [name.lower().strip() for name in reader.fieldnames if name]

            # Check for required columns
            missing_columns = []
            for required in REQUIRED_COLUMNS:
                # Check for exact match or common variations
                if required not in detected_columns:
                    # Check for variations (e.g., "match type" vs "match_type")
                    variations = [required.replace(" ", "_"), required.replace("_", " ")]
                    if not any(var in detected_columns for var in variations):
                        missing_columns.append(required)

            return detected_columns, missing_columns

    except Exception as exc:
        logger.error(f"Error validating CSV columns: {exc}")
        return [], list(REQUIRED_COLUMNS)


def validate_excel_columns(file_path: Path) -> tuple[list[str], list[str]]:
    """Validate Excel file has required columns.

    Args:
        file_path: Path to the Excel file

    Returns:
        Tuple of (detected_columns, missing_columns)
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required for Excel validation")
        return [], list(REQUIRED_COLUMNS)

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active

        # Get header row
        headers = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]

        if not headers or all(h == "" for h in headers):
            workbook.close()
            return [], list(REQUIRED_COLUMNS)

        detected_columns = [h for h in headers if h]

        # Check for required columns
        missing_columns = []
        for required in REQUIRED_COLUMNS:
            if required not in detected_columns:
                # Check for variations
                variations = [required.replace(" ", "_"), required.replace("_", " ")]
                if not any(var in detected_columns for var in variations):
                    missing_columns.append(required)

        workbook.close()
        return detected_columns, missing_columns

    except Exception as exc:
        logger.error(f"Error validating Excel columns: {exc}")
        return [], list(REQUIRED_COLUMNS)


def get_file_preview(file_path: Path, max_rows: int = 10) -> tuple[list[dict[str, Any]], int]:
    """Get preview of file contents.

    Args:
        file_path: Path to the file
        max_rows: Maximum number of rows to preview

    Returns:
        Tuple of (preview_rows, total_row_count)
    """
    file_ext = file_path.suffix.lower()

    if file_ext == ".csv":
        return _preview_csv(file_path, max_rows)
    elif file_ext in (".xlsx", ".xls"):
        return _preview_excel(file_path, max_rows)
    else:
        return [], 0


def _preview_csv(file_path: Path, max_rows: int) -> tuple[list[dict[str, Any]], int]:
    """Preview CSV file contents."""
    preview_rows = []
    total_rows = 0

    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)

            for i, row in enumerate(reader):
                total_rows += 1
                if i < max_rows:
                    preview_rows.append(row)
                elif i >= 1000:  # Stop counting after 1000 rows for performance
                    break

    except Exception as exc:
        logger.error(f"Error previewing CSV: {exc}")

    return preview_rows, total_rows


def _preview_excel(file_path: Path, max_rows: int) -> tuple[list[dict[str, Any]], int]:
    """Preview Excel file contents."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required for Excel preview")
        return [], 0

    preview_rows = []
    total_rows = 0

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [str(cell.value) if cell.value else "" for cell in sheet[1]]

        # Get data rows
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            total_rows += 1
            if i < max_rows:
                row_dict = dict(zip(headers, row))
                preview_rows.append(row_dict)
            elif i >= 1000:  # Stop counting after 1000 rows
                break

        workbook.close()

    except Exception as exc:
        logger.error(f"Error previewing Excel: {exc}")

    return preview_rows, total_rows
